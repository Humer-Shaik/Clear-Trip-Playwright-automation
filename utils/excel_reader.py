from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


@dataclass(frozen=True)
class FlightSearchData:
    test_id: str
    from_city: str
    to_city: str
    journey_type: str
    departure_days_ahead: int
    enabled: bool
    description: str

    @classmethod
    def from_row(cls, row: dict[str, Any]) -> "FlightSearchData":
        return cls(
            test_id=str(row["test_id"]).strip(),
            from_city=str(row["from_city"]).strip(),
            to_city=str(row["to_city"]).strip(),
            journey_type=str(row.get("journey_type", "One Way")).strip(),
            departure_days_ahead=int(row.get("departure_days_ahead", 7)),
            enabled=str(row.get("enabled", "Y")).strip().upper() in {"Y", "YES", "TRUE", "1"},
            description=str(row.get("description", "")).strip(),
        )


class ExcelReader:
    REQUIRED_COLUMNS = (
        "test_id",
        "from_city",
        "to_city",
        "journey_type",
        "departure_days_ahead",
        "enabled",
        "description",
    )

    def __init__(self, file_path: Path, sheet_name: str) -> None:
        self.file_path = file_path
        self.sheet_name = sheet_name

    def read_flight_search_data(self) -> list[FlightSearchData]:
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel data file not found: {self.file_path}")

        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        if self.sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{self.sheet_name}' not found. Available: {workbook.sheetnames}"
            )

        sheet = workbook[self.sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in next(rows)]

        missing = [col for col in self.REQUIRED_COLUMNS if col not in headers]
        if missing:
            raise ValueError(f"Missing required columns in Excel sheet: {missing}")

        test_data: list[FlightSearchData] = []
        for row_values in rows:
            if not row_values or all(value is None or str(value).strip() == "" for value in row_values):
                continue

            row_dict = {
                headers[index]: row_values[index]
                for index in range(len(headers))
                if headers[index]
            }
            record = FlightSearchData.from_row(row_dict)
            if record.enabled:
                test_data.append(record)

        workbook.close()
        return test_data


def create_sample_excel(file_path: Path, sheet_name: str = "FlightSearch") -> None:
    file_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    headers = list(ExcelReader.REQUIRED_COLUMNS)
    sheet.append(headers)
    sheet.append(["TC001", "Mumbai", "Delhi", "One Way", 7, "Y", "Domestic one-way search"])
    sheet.append(["TC002", "Bengaluru", "Goa", "One Way", 10, "Y", "South India route search"])
    sheet.append(["TC003", "Chennai", "Hyderabad", "One Way", 14, "Y", "Short domestic route"])
    sheet.append(["TC004", "Delhi", "Mumbai", "Round Trip", 5, "N", "Disabled sample row"])

    workbook.save(file_path)
    workbook.close()
